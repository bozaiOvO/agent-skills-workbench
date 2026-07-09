#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from string import whitespace
from typing import Iterable
from unicodedata import name as unicode_name

from openpyxl import load_workbook

DEFAULT_VOLUME_ROOT = Path('/Users/jinbo/AutomationCenter/workspace/TikTokDownloader-master/Volume')
DEFAULT_OUTPUT_ROOT = DEFAULT_VOLUME_ROOT / '整理输出_按年份热度'
TEXT_ENCODINGS = ('utf-8', 'utf-8-sig', 'gb18030')
TEMP_FILE_PREFIXES = ('._',)
TEMP_FILE_MARKERS = ('.~lock',)
CONTROL_CHARACTERS = re.compile(r'[\x00-\x1F\x7F]')
REQUIRED_COLUMNS = {
    '作品类型',
    '作品ID',
    '作品描述',
    '作品话题',
    '视频时长',
    '发布时间',
    '账号昵称',
    '点赞数量',
    '评论数量',
    '收藏数量',
    '分享数量',
}
FILENAME_RULE = {
    '/': '',
    '\\': '',
    '|': '',
    '<': '',
    '>': '',
    '"': '',
    '?': '',
    ':': '',
    '*': '',
    '\x00': '',
}
FILENAME_RULE |= {item: '' for item in whitespace[1:]}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Rank transcript texts by yearly hotness using Volume/Data workbooks.')
    parser.add_argument('--volume-root', type=Path, default=DEFAULT_VOLUME_ROOT)
    parser.add_argument('--output-root', type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument('--rank-width', type=int, default=3)
    parser.add_argument('--stem', dest='stems', action='append', default=[])
    parser.add_argument('--earliest', default='', help='Only include works published on/after this date')
    parser.add_argument('--latest', default='', help='Only include works published on/before this date')
    return parser.parse_args()


def is_temporary_file(path: Path) -> bool:
    return path.name.startswith(TEMP_FILE_PREFIXES) or any(marker in path.name for marker in TEMP_FILE_MARKERS)


def is_chinese_char(char: str) -> bool:
    return 'CJK' in unicode_name(char, '')


def truncate_string(text: str, length: int) -> str:
    count = 0
    result = []
    for char in text:
        count += 2 if is_chinese_char(char) else 1
        if count > length:
            break
        result.append(char)
    return ''.join(result)


def beautify_string(text: str, length: int) -> str:
    count = 0
    for char in text:
        count += 2 if is_chinese_char(char) else 1
        if count > length:
            break
    else:
        return text
    half = length // 2
    start = truncate_string(text, half)
    end = truncate_string(text[::-1], half)[::-1]
    return f'{start}...{end}'


def clear_spaces(text: str) -> str:
    return ' '.join(text.split())


def filter_name(text: str, default: str = '') -> str:
    cleaned = text.replace(':', '.')
    cleaned = CONTROL_CHARACTERS.sub('', cleaned)
    for old, new in FILENAME_RULE.items():
        cleaned = cleaned.replace(old, new)
    cleaned = clear_spaces(cleaned).strip().strip('.')
    return cleaned or default


def read_text_with_fallback(path: Path) -> str:
    last_error: Exception | None = None
    for encoding in TEXT_ENCODINGS:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as error:
            last_error = error
    if last_error is not None:
        raise last_error
    return path.read_text(encoding='utf-8')


def safe_int(value) -> int:
    if value in (None, ''):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(',', '')
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def normalize_datetime_text(value) -> str:
    parsed = parse_datetime_value(value)
    if parsed is not None:
        return parsed.strftime('%Y-%m-%d %H:%M:%S')
    text = str(value or '').strip()
    if not text:
        return ''
    return text[:19] if len(text) >= 19 else text


def parse_datetime_value(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = str(value or '').strip()
    if not text:
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M', '%Y-%m-%d', '%Y/%m/%d'):
        try:
            parsed = datetime.strptime(text, fmt)
            if 'H' not in fmt:
                return parsed.replace(hour=0, minute=0, second=0)
            return parsed
        except ValueError:
            continue
    return None


def extract_year(publish_time: str) -> str:
    match = re.search(r'(20\d{2})', publish_time)
    return match.group(1) if match else '未知年份'


def extract_tags(topic: str, description: str) -> str:
    topic_text = clear_spaces(str(topic or ''))
    if topic_text:
        return topic_text
    tags = re.findall(r'#([^#\s]+)', str(description or ''))
    return ' '.join(tags)


def extract_title(description: str) -> str:
    text = clear_spaces(str(description or ''))
    if not text:
        return ''
    if '#' in text:
        return text.split('#', 1)[0].strip()
    return text


def workbook_rows(path: Path) -> tuple[list[str], Iterable[tuple]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    header = [str(cell or '').strip() for cell in next(iterator)]
    missing = sorted(REQUIRED_COLUMNS - set(header))
    if missing:
        raise ValueError(f'{path.name} missing required columns: {", ".join(missing)}')
    return header, iterator


def row_to_dict(header: list[str], row: tuple) -> dict[str, object]:
    result: dict[str, object] = {}
    for index, key in enumerate(header):
        result[key] = row[index] if index < len(row) else None
    return result


def latest_row_key(row_data: dict[str, object]) -> str:
    work_id = str(row_data.get('作品ID') or '').strip()
    if work_id:
        return f'id:{work_id}'
    return f"stem:{expected_transcript_stem(row_data)}"


def is_newer_collection_row(candidate: dict[str, object], current: dict[str, object]) -> bool:
    candidate_collected = parse_datetime_value(candidate.get('采集时间'))
    current_collected = parse_datetime_value(current.get('采集时间'))
    if candidate_collected is not None or current_collected is not None:
        if current_collected is None:
            return True
        if candidate_collected is None:
            return False
        if candidate_collected != current_collected:
            return candidate_collected > current_collected
    return safe_int(candidate.get('__row_number')) > safe_int(current.get('__row_number'))


def latest_workbook_rows(header: list[str], rows: Iterable[tuple]) -> list[dict[str, object]]:
    latest: dict[str, dict[str, object]] = {}
    variants: dict[str, list[dict[str, object]]] = {}
    for row_number, row in enumerate(rows, start=2):
        row_data = row_to_dict(header, row)
        row_data['__row_number'] = row_number
        key = latest_row_key(row_data)
        variants.setdefault(key, []).append(row_data)
        current = latest.get(key)
        if current is None or is_newer_collection_row(row_data, current):
            latest[key] = row_data
    for key, row_data in latest.items():
        row_data['__match_rows'] = variants.get(key, [])
    return sorted(latest.values(), key=lambda item: safe_int(item.get('__row_number')))


def match_row_variants(row_data: dict[str, object]) -> list[dict[str, object]]:
    raw_variants = row_data.get('__match_rows')
    variants = raw_variants if isinstance(raw_variants, list) else []
    ordered: list[dict[str, object]] = []
    seen: set[int] = set()
    for candidate in [row_data, *sorted(variants, key=lambda item: safe_int(item.get('__row_number')), reverse=True)]:
        row_number = safe_int(candidate.get('__row_number'))
        if row_number in seen:
            continue
        seen.add(row_number)
        ordered.append(candidate)
    return ordered


def expected_transcript_stem(row: dict[str, object]) -> str:
    publish_time = normalize_datetime_text(row.get('发布时间'))
    desc = beautify_string(str(row.get('作品描述') or ''), 64)
    pieces = [publish_time, str(row.get('作品类型') or ''), str(row.get('账号昵称') or ''), desc]
    stem = filter_name('-'.join(pieces), default=str(row.get('作品ID') or ''))
    return beautify_string(stem, 128)


def transcript_prefix(row: dict[str, object]) -> str:
    publish_time = normalize_datetime_text(row.get('发布时间'))
    pieces = [publish_time, str(row.get('作品类型') or ''), str(row.get('账号昵称') or '')]
    return filter_name('-'.join(pieces), default='')


def transcript_map(folder: Path) -> dict[str, Path]:
    result: dict[str, Path] = {}
    for path in folder.iterdir():
        if not path.is_file() or is_temporary_file(path):
            continue
        if path.suffix.lower() != '.txt':
            continue
        result[path.stem] = path
    return result


def score_row(row: dict[str, object]) -> int:
    comments = safe_int(row.get('评论数量'))
    favorites = safe_int(row.get('收藏数量'))
    shares = safe_int(row.get('分享数量'))
    likes = safe_int(row.get('点赞数量'))
    return comments * 4 + favorites * 3 + shares * 2 + likes


def infer_blogger_name(folder_name: str) -> str:
    match = re.match(r'^[A-Z]+\d+_(.+?)_(发布作品|喜欢作品|收藏作品|收藏夹作品|合集作品)$', folder_name)
    if match:
        name, kind = match.groups()
        if kind == '收藏夹作品':
            return f'收藏夹/{name}'
        if kind == '喜欢作品':
            return f'点赞/{name}'
        if kind == '收藏作品':
            return f'收藏/{name}'
        return name
    return folder_name


def should_use_account_nickname(folder_name: str) -> bool:
    return folder_name.startswith('UID') and folder_name.endswith('_发布作品')


def output_dir_for_name(output_root: Path, name: str) -> Path:
    parts = [filter_name(part, part) for part in str(name).split('/') if part.strip()]
    if not parts:
        parts = [filter_name(name, name)]
    return output_root.joinpath(*parts)


BODY_DELIMITER = '=======下为正文============'
METADATA_PREFIXES = (
    '评论：',
    '收藏：',
    '分享：',
    '点赞：',
    '视频时长：',
    '发布时间：',
    '综合分：',
    '标题：',
    '标签：',
)


def strip_metadata_header(text: str) -> str:
    if BODY_DELIMITER not in text:
        return text.rstrip()
    header, body = text.split(BODY_DELIMITER, 1)
    header_lines = [line.strip() for line in header.splitlines() if line.strip()]
    if any(line.startswith(METADATA_PREFIXES) for line in header_lines):
        return body.lstrip('\r\n').rstrip()
    return text.rstrip()


def render_output(item: dict[str, object]) -> str:
    body = strip_metadata_header(str(item['body']))
    lines = [
        f"评论：{item['comments']}",
        f"收藏：{item['favorites']}",
        f"分享：{item['shares']}",
        f"点赞：{item['likes']}",
        f"视频时长：{item['duration']}",
        f"发布时间：{item['publish_time']}",
        f"综合分：{item['score']}",
        f"标题：{item['title']}",
        f"标签：{item['tags']}",
        BODY_DELIMITER,
        body,
    ]
    return '\n'.join(lines).rstrip() + '\n'


def collect_items(
    workbook_path: Path,
    blogger_folder: Path,
    earliest: datetime | None = None,
    latest: datetime | None = None,
) -> tuple[str, dict[str, list[dict[str, object]]], dict[str, int]]:
    header, rows = workbook_rows(workbook_path)
    row_datas = latest_workbook_rows(header, rows)
    txt_files = transcript_map(blogger_folder)
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    stats = {'matched': 0, 'unmatched': 0}
    blogger_name = infer_blogger_name(blogger_folder.name)
    use_account_nickname = should_use_account_nickname(blogger_folder.name)
    for row_data in row_datas:
        if use_account_nickname:
            blogger_name = str(row_data.get('账号昵称') or blogger_name).strip() or blogger_name
        publish_dt = parse_datetime_value(row_data.get('发布时间'))
        if earliest is not None and (publish_dt is None or publish_dt < earliest):
            continue
        if latest is not None and (publish_dt is None or publish_dt > latest):
            continue
        stem = expected_transcript_stem(row_data)
        transcript = None
        for match_row in match_row_variants(row_data):
            match_stem = expected_transcript_stem(match_row)
            transcript = txt_files.get(match_stem)
            if transcript is not None:
                break
            prefix = transcript_prefix(match_row)
            candidates = [path for key, path in txt_files.items() if prefix and key.startswith(prefix)]
            if len(candidates) == 1:
                transcript = candidates[0]
                break
        if transcript is None:
            stats['unmatched'] += 1
            continue
        work_id = str(row_data.get('作品ID') or '').strip()
        publish_time = normalize_datetime_text(row_data.get('发布时间'))
        grouped[extract_year(publish_time)].append(
            {
                'stem': stem,
                'body': read_text_with_fallback(transcript),
                'comments': safe_int(row_data.get('评论数量')),
                'favorites': safe_int(row_data.get('收藏数量')),
                'shares': safe_int(row_data.get('分享数量')),
                'likes': safe_int(row_data.get('点赞数量')),
                'duration': str(row_data.get('视频时长') or ''),
                'publish_time': publish_time,
                'publish_sort': publish_time,
                'score': score_row(row_data),
                'title': extract_title(row_data.get('作品描述')),
                'tags': extract_tags(row_data.get('作品话题'), row_data.get('作品描述')),
                'work_id': work_id,
            }
        )
        stats['matched'] += 1
    return blogger_name, grouped, stats


def write_output(output_root: Path, blogger_name: str, grouped_items: dict[str, list[dict[str, object]]], rank_width: int) -> int:
    written = 0
    blogger_dir = output_dir_for_name(output_root, blogger_name)
    for year, items in sorted(grouped_items.items()):
        year_dir = blogger_dir / str(year)
        year_dir.mkdir(parents=True, exist_ok=True)
        for old_path in year_dir.glob('top*.txt'):
            if old_path.is_file() and not is_temporary_file(old_path):
                old_path.unlink()
        items.sort(
            key=lambda item: (
                -int(item['score']),
                -int(item['comments']),
                -int(item['favorites']),
                -int(item['shares']),
                -int(item['likes']),
                str(item['publish_sort']),
                str(item['stem']),
            )
        )
        width = max(rank_width, len(str(len(items))))
        for rank, item in enumerate(items, start=1):
            prefix = f'top{rank:0{width}d}'
            filename = filter_name(f"{prefix}_{item['stem']}", default=f"{prefix}_{item['work_id']}")
            (year_dir / f'{filename}.txt').write_text(render_output(item), encoding='utf-8')
            written += 1
    return written


def main() -> int:
    args = parse_args()
    volume_root = args.volume_root.expanduser().resolve()
    data_root = volume_root / 'Data'
    output_root = args.output_root.expanduser().resolve()
    stems = set(args.stems)
    earliest = parse_datetime_value(args.earliest)
    latest = parse_datetime_value(args.latest)
    if args.earliest and earliest is None:
        raise SystemExit(f'Invalid --earliest date: {args.earliest}')
    if args.latest and latest is None:
        raise SystemExit(f'Invalid --latest date: {args.latest}')
    if not volume_root.is_dir():
        raise SystemExit(f'Volume root not found: {volume_root}')
    if not data_root.is_dir():
        raise SystemExit(f'Data root not found: {data_root}')
    output_root.mkdir(parents=True, exist_ok=True)
    workbook_paths = sorted(
        path
        for path in data_root.glob('*.xlsx')
        if path.is_file() and not is_temporary_file(path) and (not stems or path.stem in stems)
        and re.match(r'^(UID|CID|MID)\d+_', path.stem)
    )
    total_written = 0
    total_matched = 0
    total_unmatched = 0
    for workbook_path in workbook_paths:
        blogger_folder = volume_root / workbook_path.stem
        if not blogger_folder.is_dir():
            print(f'SKIP no-folder {workbook_path.name}')
            continue
        blogger_name, grouped_items, stats = collect_items(workbook_path, blogger_folder, earliest=earliest, latest=latest)
        if not grouped_items:
            print(f'SKIP no-transcripts {workbook_path.name}')
            continue
        written = write_output(output_root, blogger_name, grouped_items, args.rank_width)
        total_written += written
        total_matched += stats['matched']
        total_unmatched += stats['unmatched']
        year_counts = ', '.join(f'{year}:{len(items)}' for year, items in sorted(grouped_items.items()))
        print(f"OK {blogger_name} matched={stats['matched']} unmatched={stats['unmatched']} written={written} years=[{year_counts}]")
    print(f'SUMMARY written={total_written} matched={total_matched} unmatched={total_unmatched} output={output_root}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
